

import argparse
import sys
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import (
    friedmanchisquare,
    wilcoxon,
    chi2_contingency,
    binomtest,
    norm,
)

# ===========================================================
# CONFIG
# ===========================================================

RANDOM_SEED = 42
ALPHA = 0.05
PROVIDERS = ["OpenAI", "Claude", "Gemini", "Deepseek"]
SCHEMA_MODES = ["zero_shot", "one_shot", "two_shot"]

DATA_DIR = Path(__file__).parent / "data"

# Data file paths (adjust to match your repo)
GT_F1_PATH = DATA_DIR / "groundtruth_f1.xlsx"  # ground-truth F1 per cell
CMR_PATH = DATA_DIR / "cross_model_review.xlsm"  # cross-model review tracker
QA_OUTCOMES_PATH = DATA_DIR / "qa_outcomes.csv"  # Q&A 3x4 contingency
AUDIT_PATH = DATA_DIR / "audit_n316.csv"  # manual audit results

# ===========================================================
# UTILITY FUNCTIONS
# ===========================================================


def kendalls_w(chi2: float, n: int, k: int) -> float:
    """Kendall's W from Friedman chi-square.

    W = chi^2 / [n * (k - 1)]
    where n is the number of papers (blocks) and k is the number of
    methods being compared.
    """
    return chi2 / (n * (k - 1))


def cramers_v(chi2: float, n: int, r: int, c: int) -> float:
    """Cramer's V for a chi-square test of independence.

    V = sqrt(chi^2 / [n * min(r-1, c-1)])
    """
    return float(np.sqrt(chi2 / (n * min(r - 1, c - 1))))


def wilson_ci(successes: int, n: int, alpha: float = 0.05) -> tuple:
    """Wilson 95% confidence interval for a proportion.

    More accurate than the normal approximation for small n or
    extreme proportions; recommended by Brown, Cai & DasGupta (2001).
    """
    if n == 0:
        return (0.0, 0.0)
    z = norm.ppf(1 - alpha / 2)
    p = successes / n
    denom = 1 + z**2 / n
    center = (p + z**2 / (2 * n)) / denom
    margin = z * np.sqrt(p * (1 - p) / n + z**2 / (4 * n**2)) / denom
    return (center - margin, center + margin)


def holm_bonferroni(pairs: list) -> list:
    """Apply Holm-Bonferroni correction to a list of (name, raw_p) tuples.

    Returns list of (name, raw_p, holm_adjusted_p) preserving original
    input order.
    """
    indexed = list(enumerate(pairs))
    sorted_pairs = sorted(indexed, key=lambda x: x[1][1])
    m = len(sorted_pairs)
    adjusted = [None] * m
    prev_adj = 0.0
    for rank, (orig_idx, (name, p)) in enumerate(sorted_pairs):
        adj = min(1.0, max(prev_adj, p * (m - rank)))
        prev_adj = adj
        adjusted[orig_idx] = (name, p, adj)
    return adjusted


def pairwise_wilcoxon(pivot: pd.DataFrame, groups: list) -> list:
    """Pairwise Wilcoxon signed-rank tests with Holm-Bonferroni correction.

    pivot: DataFrame with one column per group, rows are paired observations
    groups: list of column names to compare pairwise
    """
    pairs = []
    for a, b in combinations(groups, 2):
        try:
            _, p = wilcoxon(pivot[a], pivot[b], zero_method="wilcox")
        except ValueError:
            # all differences are zero - no detectable effect
            p = 1.0
        pairs.append((f"{a} vs {b}", p))
    return holm_bonferroni(pairs)


def print_section_header(label: str) -> None:
    print("\n" + "=" * 70)
    print(label)
    print("=" * 70)


def print_friedman_result(chi2: float, p: float, W: float, k: int, n: int) -> None:
    print(f"  Friedman: chi2({k-1}) = {chi2:.3f}, p = {p:.4f}, Kendall's W = {W:.3f}")
    print(f"  n papers = {n}, k methods = {k}")


# ===========================================================
# DATA LOADERS
# ===========================================================


def load_groundtruth_f1() -> pd.DataFrame:
    """Load ground-truth F1 scores from groundtruth_f1.xlsx.

    Expected long-format columns: paper, provider, schema_mode, f1.
    Adapt this loader if your stored format differs.
    """
    if not GT_F1_PATH.exists():
        raise FileNotFoundError(f"Ground-truth F1 file not found: {GT_F1_PATH}")
    df = pd.read_excel(GT_F1_PATH)
    required = {"paper", "provider", "schema_mode", "f1"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"groundtruth_f1.xlsx missing columns: {missing}")
    return df


def load_cross_model_review() -> pd.DataFrame:
    """Load cross-model review F1 from the multi-sheet tracker xlsm.

    Sheet structure: one sheet per paper named 'Paper 1', 'Paper 2', etc.
    Within each sheet, four blocks:
      - Rows 19-25: OpenAI extractor (header row 19, F1 in row 25)
      - Rows 27-33: Claude  extractor (header row 27, F1 in row 33)
      - Rows 35-41: Gemini  extractor (header row 35, F1 in row 41)
      - Rows 43-49: DeepSeek extractor (header row 43, F1 in row 49)
    Header row: [extractor_name, marker1, marker2, marker3] in columns A-D.
    F1 row: [F1_label, F1_marker1, F1_marker2, F1_marker3] in columns A-D.

    Blocks identical to the corresponding block in Paper 1 are treated
    as unfilled template duplicates and skipped.
    """
    import openpyxl

    if not CMR_PATH.exists():
        raise FileNotFoundError(f"Cross-model review file not found: {CMR_PATH}")

    wb = openpyxl.load_workbook(CMR_PATH, data_only=True)
    blocks = [(19, 25, "OpenAI"), (27, 33, "Claude"), (35, 41, "Gemini"), (43, 49, "Deepseek")]

    # Reference signatures from Paper 1 (used to detect template duplicates)
    ref_sigs = {}
    sheet1 = wb["Paper 1"]
    for hdr_row, _, name in blocks:
        ref_sigs[name] = tuple(sheet1.cell(row=hdr_row + 1, column=c).value for c in range(2, 5))

    records = []
    for i in range(1, 21):
        sheet_name = f"Paper {i}"
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        for hdr_row, f1_row, ext_name in blocks:
            tp = tuple(sheet.cell(row=hdr_row + 1, column=c).value for c in range(2, 5))
            # Skip unfilled template duplicates
            if i != 1 and tp == ref_sigs[ext_name]:
                continue
            for col in range(2, 5):
                marker = sheet.cell(row=hdr_row, column=col).value
                f1 = sheet.cell(row=f1_row, column=col).value
                if marker is None or f1 is None:
                    continue
                records.append({
                    "paper": i,
                    "extractor": ext_name,
                    "marker": str(marker).strip(),
                    "f1": float(f1),
                })
    return pd.DataFrame(records)


def load_qa_outcomes() -> pd.DataFrame:
    """Load Q&A pipeline outcomes from qa_outcomes.csv.

    Expected long-format columns: paper, reader, schema_mode, outcome
    where outcome in {agree, disagree, kb_silent, error}.
    """
    if not QA_OUTCOMES_PATH.exists():
        raise FileNotFoundError(f"Q&A outcomes file not found: {QA_OUTCOMES_PATH}")
    return pd.read_csv(QA_OUTCOMES_PATH)


def load_audit() -> pd.DataFrame:
    """Load manual audit results from audit_n316.csv.

    Expected columns: paper, reader, question_id, path_a_verdict, path_b_verdict
    where verdicts are in {Y, N, P, X, -}.
    """
    if not AUDIT_PATH.exists():
        raise FileNotFoundError(f"Audit file not found: {AUDIT_PATH}")
    return pd.read_csv(AUDIT_PATH)


# ===========================================================
# H1: GROUND-TRUTH F1 DIFFERS BY PROVIDER
# Methodology: Section 3.6 | Hypothesis: 3.10.2 | Results: Section 5.1
# ===========================================================


def test_H1_groundtruth_f1() -> dict:
    """H1: Per-paper ground-truth F1 differs across the four providers
    within at least one schema mode. Tested by Friedman, one test per
    schema mode (3 tests total)."""

    print_section_header("H1: GROUND-TRUTH F1 BY PROVIDER (Friedman per schema mode)")

    df = load_groundtruth_f1()
    results = {}

    for schema in SCHEMA_MODES:
        sub = df[df["schema_mode"] == schema]
        pivot = sub.pivot(index="paper", columns="provider", values="f1")[PROVIDERS]
        n, k = len(pivot), 4

        chi2, p = friedmanchisquare(*[pivot[col].values for col in PROVIDERS])
        W = kendalls_w(chi2, n, k)

        print(f"\n{schema}:")
        print_friedman_result(chi2, p, W, k, n)

        print(f"  Mean F1 by provider:")
        for col in PROVIDERS:
            print(f"    {col:10s} {pivot[col].mean():.3f}")

        if p < ALPHA:
            print(f"  Post-hoc Wilcoxon (Holm-corrected):")
            for name, raw, adj in pairwise_wilcoxon(pivot, PROVIDERS):
                sig = " *" if adj < ALPHA else ""
                print(f"    {name:25s} raw={raw:.4f} holm={adj:.4f}{sig}")

        results[schema] = {"chi2": chi2, "p": p, "W": W, "n": n}

    return results


# ===========================================================
# H2: WITHIN-PROVIDER SCHEMA EFFECT
# Methodology: Section 3.6 | Hypothesis: 3.10.2 | Results: Section 5.1
# ===========================================================


def test_H2_schema_effect() -> dict:
    """H2: Per-paper ground-truth F1 differs across the three schema modes
    for at least one provider. Tested by Friedman per provider (4 tests)."""

    print_section_header("H2: SCHEMA EFFECT WITHIN EACH PROVIDER (Friedman per provider)")

    df = load_groundtruth_f1()
    results = {}

    for provider in PROVIDERS:
        sub = df[df["provider"] == provider]
        pivot = sub.pivot(index="paper", columns="schema_mode", values="f1")[SCHEMA_MODES]
        n, k = len(pivot), 3

        chi2, p = friedmanchisquare(*[pivot[col].values for col in SCHEMA_MODES])
        W = kendalls_w(chi2, n, k)

        print(f"\n{provider}:")
        print_friedman_result(chi2, p, W, k, n)

        if p < ALPHA:
            print(f"  Post-hoc Wilcoxon (Holm-corrected):")
            for name, raw, adj in pairwise_wilcoxon(pivot, SCHEMA_MODES):
                sig = " *" if adj < ALPHA else ""
                print(f"    {name:25s} raw={raw:.4f} holm={adj:.4f}{sig}")

        results[provider] = {"chi2": chi2, "p": p, "W": W, "n": n}

    return results


# ===========================================================
# H3: CROSS-MODEL MARKER DISAGREEMENT
# Methodology: Section 3.7 | Hypothesis: 3.10.2 | Results: Section 5.X
# ===========================================================


def test_H3_marker_disagreement() -> dict:
    """H3: Mean F1 given differs across the four LLM markers. The core
    LLM-as-evaluator question. Two analyses are run:

    Analysis 1: marker generosity (one F1 mean per (paper, marker) cell)
    Analysis 2: marker consensus on extractor quality (one F1 mean per
                (paper, extractor) cell, averaging across markers)
    """

    print_section_header("H3: CROSS-MODEL REVIEW (marker disagreement and consensus)")

    df = load_cross_model_review()
    results = {}

    # --- Analysis 1: marker generosity -------------------------------
    print("\nANALYSIS 1: Marker generosity")
    print("  (For each paper, mean F1 each marker assigned across the")
    print("   extractions they scored. Tests whether markers are differently")
    print("   lenient/strict in evaluating peer extractions.)")

    marker_means = df.groupby(["paper", "marker"])["f1"].mean().unstack()[PROVIDERS]
    complete = marker_means.dropna()
    n, k = len(complete), 4

    chi2, p = friedmanchisquare(*[complete[col].values for col in PROVIDERS])
    W = kendalls_w(chi2, n, k)
    print_friedman_result(chi2, p, W, k, n)

    print("\n  Mean F1 given by marker:")
    for col in PROVIDERS:
        print(f"    {col:10s} {complete[col].mean():.3f}")

    print("  Mean rank (1 = most lenient):")
    ranks = complete.rank(axis=1, ascending=False)
    for col in PROVIDERS:
        print(f"    {col:10s} {ranks[col].mean():.2f}")

    if p < ALPHA:
        print("\n  Post-hoc Wilcoxon (Holm-corrected):")
        for name, raw, adj in pairwise_wilcoxon(complete, PROVIDERS):
            sig = " *" if adj < ALPHA else ""
            print(f"    {name:25s} raw={raw:.4f} holm={adj:.4f}{sig}")

    results["analysis_1_generosity"] = {"chi2": chi2, "p": p, "W": W, "n": n}

    # --- Analysis 2: extractor consensus -----------------------------
    print("\n\nANALYSIS 2: Marker consensus on extraction quality")
    print("  (For each paper, mean F1 each extractor received averaged across")
    print("   the three markers. Tests whether the marker consensus differentiates")
    print("   extractors. Reported descriptively if Analysis 1 rejects.)")

    extractor_means = df.groupby(["paper", "extractor"])["f1"].mean().unstack()[PROVIDERS]
    complete_e = extractor_means.dropna()
    n_e, k_e = len(complete_e), 4

    chi2_e, p_e = friedmanchisquare(*[complete_e[col].values for col in PROVIDERS])
    W_e = kendalls_w(chi2_e, n_e, k_e)
    print_friedman_result(chi2_e, p_e, W_e, k_e, n_e)

    print("\n  Mean consensus F1 by extractor:")
    for col in PROVIDERS:
        print(f"    {col:10s} {complete_e[col].mean():.3f}")

    results["analysis_2_consensus"] = {"chi2": chi2_e, "p": p_e, "W": W_e, "n": n_e}

    return results


# ===========================================================
# H4: Q&A SCHEMA-MODE x OUTCOME INDEPENDENCE
# Methodology: Section 3.8 | Hypothesis: 3.10.2 | Results: Section 5.2
# ===========================================================


def test_H4_qa_schema_outcome() -> dict:
    """H4: The Q&A outcome distribution depends on schema mode.
    Tested by 3 x 4 Pearson chi-square on (schema_mode x outcome)."""

    print_section_header("H4: Q&A OUTCOME DEPENDS ON SCHEMA MODE (3 x 4 chi-square)")

    df = load_qa_outcomes()

    # Build contingency table: rows = schema_mode, columns = outcome
    outcomes = ["agree", "disagree", "kb_silent", "error"]
    table = pd.crosstab(df["schema_mode"], df["outcome"])
    table = table.reindex(index=SCHEMA_MODES, columns=outcomes, fill_value=0)

    chi2, p, dof, expected = chi2_contingency(table.values)
    n = table.values.sum()
    V = cramers_v(chi2, n, *table.shape)

    print(f"\n  chi2({dof}) = {chi2:.3f}, p = {p:.4g}, Cramer's V = {V:.3f}")
    print(f"  n questions = {n}")

    print("\n  Contingency table (counts):")
    print(table)

    print("\n  Row proportions (within each schema mode):")
    prop = table.div(table.sum(axis=1), axis=0)
    print(prop.round(3))

    # Wilson 95% CIs on agreement rate per schema mode
    print("\n  Wilson 95% CI on agreement rate per schema mode:")
    for schema in SCHEMA_MODES:
        successes = int(table.loc[schema, "agree"])
        n_schema = int(table.loc[schema].sum())
        lo, hi = wilson_ci(successes, n_schema)
        rate = successes / n_schema if n_schema > 0 else 0.0
        print(f"    {schema:12s} {rate:.3f}  [{lo:.3f}, {hi:.3f}]  n={n_schema}")

    return {"chi2": chi2, "p": p, "V": V, "n": n, "df": dof}


# ===========================================================
# H5: PATH A VS PATH B HEAD-TO-HEAD
# Methodology: Section 3.8.6 | Hypothesis: 3.10.2 | Results: Section 5.3
# ===========================================================


def test_H5_path_a_vs_b() -> dict:
    """H5: Path A and Path B differ on jointly-attempted questions with
    one direction systematically dominating. Tested by McNemar's exact test
    pooled across the four reader batches."""

    print_section_header("H5: PATH A VS PATH B (McNemar's exact test, pooled)")

    df = load_audit()

    # Restrict to jointly-attempted questions (both paths produced an answer)
    attempts = df[(df["path_a_verdict"] != "-") & (df["path_b_verdict"] != "-")]

    # Treat Y as correct, {N, P} as incorrect
    a_correct = attempts["path_a_verdict"] == "Y"
    b_correct = attempts["path_b_verdict"] == "Y"

    # 2x2 contingency: rows = Path A correct/incorrect, cols = Path B correct/incorrect
    both = ((a_correct) & (b_correct)).sum()
    a_only = ((a_correct) & (~b_correct)).sum()
    b_only = ((~a_correct) & (b_correct)).sum()
    neither = ((~a_correct) & (~b_correct)).sum()

    print(f"\n  Jointly-attempted questions: n = {len(attempts)}")
    print(f"  Both correct:         {both}")
    print(f"  Path A only correct:  {a_only}  (discordant: A wins)")
    print(f"  Path B only correct:  {b_only}  (discordant: B wins)")
    print(f"  Neither correct:      {neither}")

    # McNemar's exact: binomial test on discordant counts
    discordant = a_only + b_only
    if discordant == 0:
        print("  No discordant pairs; McNemar's not applicable.")
        return {"discordant": 0}

    result = binomtest(a_only, discordant, p=0.5, alternative="two-sided")
    p = result.pvalue
    odds_ratio = (a_only / discordant) / (b_only / discordant) if b_only else float("inf")

    print(f"\n  McNemar's exact test (binomial on discordant pairs):")
    print(f"    A wins / total discordant = {a_only} / {discordant}")
    print(f"    p = {p:.4g}, odds ratio (A:B) = {odds_ratio:.2f}")

    return {"a_only": a_only, "b_only": b_only, "p": p, "odds_ratio": odds_ratio}


# ===========================================================
# H6 and H7: READER INDEPENDENCE CHECKS
# Methodology: Sections 3.8.3 and 3.8.4 | Hypothesis: 3.10.2 | Results: Section 5.4
# ===========================================================


def test_H6_H7_reader_independence() -> dict:
    """H6, H7: Path A and Path B precision do not differ across the four
    readers. Tested by 2 x 4 chi-square (correct vs incorrect by reader).
    Failure to reject is the predicted outcome. H7 (Path B) is a
    methodological isolation check because Path B does not invoke the
    reader provider at any stage."""

    print_section_header("H6, H7: READER INDEPENDENCE (Path A and Path B)")

    df = load_audit()
    results = {}

    for path_name, verdict_col in [("Path A", "path_a_verdict"), ("Path B", "path_b_verdict")]:
        attempts = df[df[verdict_col] != "-"]
        attempts = attempts.copy()
        attempts["correct"] = (attempts[verdict_col] == "Y").astype(int)

        table = pd.crosstab(attempts["reader"], attempts["correct"])
        table = table.reindex(index=PROVIDERS, columns=[0, 1], fill_value=0)

        chi2, p, dof, _ = chi2_contingency(table.values)
        n = table.values.sum()
        V = cramers_v(chi2, n, *table.shape)

        print(f"\n  {path_name}:")
        print(f"    chi2({dof}) = {chi2:.3f}, p = {p:.4g}, Cramer's V = {V:.3f}")
        print(f"    n attempted = {n}")
        if p > ALPHA:
            print(f"    Failed to reject null (predicted outcome).")
        else:
            print(f"    Rejected null - reader effect detected, p < {ALPHA}.")
        results[path_name] = {"chi2": chi2, "p": p, "V": V, "n": n}

    return results


# ===========================================================
# BOOTSTRAP: TWO-SHOT VS ONE-SHOT AGREEMENT RATIO
# Methodology: Section 3.8.6 | Results: Section 5.2 (headline)
# ===========================================================


def bootstrap_agreement_ratio(n_resamples: int = 10_000) -> dict:
    """Bootstrap 95% CI on the two-shot vs one-shot agreement-rate ratio.

    Agreement rate = agree / (agree + disagree) among answer-attempted
    questions (kb_silent and error excluded so the schema effect is
    isolated from reader-level abstention behaviour).

    Procedure: independently resample with replacement from each schema's
    answer-attempted subset 10^4 times; compute the ratio of resampled
    means; report 2.5th and 97.5th percentiles as the 95% CI.
    """

    print_section_header(
        "BOOTSTRAP: Two-shot / one-shot agreement-rate ratio (10^4 resamples)"
    )

    df = load_qa_outcomes()

    def get_attempted(schema):
        # Binary array: 1 if agree, 0 if disagree, for answer-attempted only
        sub = df[(df["schema_mode"] == schema) & (df["outcome"].isin(["agree", "disagree"]))]
        return (sub["outcome"] == "agree").astype(int).values

    two = get_attempted("two_shot")
    one = get_attempted("one_shot")

    print(f"\n  Two-shot answer-attempted: n = {len(two)}")
    print(f"    agree = {two.sum()}, disagree = {len(two) - two.sum()}")
    print(f"    raw rate = {two.mean():.3f}")
    print(f"  One-shot answer-attempted: n = {len(one)}")
    print(f"    agree = {one.sum()}, disagree = {len(one) - one.sum()}")
    print(f"    raw rate = {one.mean():.3f}")
    print(f"  Point estimate ratio: {two.mean() / one.mean():.2f}x")

    # Bootstrap
    rng = np.random.default_rng(RANDOM_SEED)
    ratios = []
    for _ in range(n_resamples):
        t = rng.choice(two, size=len(two), replace=True)
        o = rng.choice(one, size=len(one), replace=True)
        if o.mean() > 0:
            ratios.append(t.mean() / o.mean())
    lo, hi = np.percentile(ratios, [2.5, 97.5])

    print(f"\n  Bootstrap 95% CI: [{lo:.2f}x, {hi:.2f}x]")
    print(f"  Excludes 1: {'YES' if lo > 1 else 'NO'} (lower bound > 1)")

    return {
        "two_shot_rate": float(two.mean()),
        "one_shot_rate": float(one.mean()),
        "ratio": float(two.mean() / one.mean()),
        "ci_lower": float(lo),
        "ci_upper": float(hi),
        "n_resamples": n_resamples,
    }


# ===========================================================
# MAIN
# ===========================================================


HYPOTHESIS_TESTS = {
    "H1": test_H1_groundtruth_f1,
    "H2": test_H2_schema_effect,
    "H3": test_H3_marker_disagreement,
    "H4": test_H4_qa_schema_outcome,
    "H5": test_H5_path_a_vs_b,
    "H6": test_H6_H7_reader_independence,  # H6 and H7 are run together
    "H7": test_H6_H7_reader_independence,
    "bootstrap": bootstrap_agreement_ratio,
}


def main():
    parser = argparse.ArgumentParser(description="Project 15 statistical analysis")
    parser.add_argument("--test", choices=list(HYPOTHESIS_TESTS.keys()),
                        help="Run a specific hypothesis test only")
    parser.add_argument("--bootstrap-only", action="store_true",
                        help="Run only the bootstrap agreement-ratio CI")
    args = parser.parse_args()

    np.random.seed(RANDOM_SEED)
    print(f"Random seed = {RANDOM_SEED}")
    print(f"Significance threshold alpha = {ALPHA}")
    print(f"Data directory: {DATA_DIR.resolve()}")

    if args.bootstrap_only or args.test == "bootstrap":
        bootstrap_agreement_ratio()
        return

    if args.test:
        HYPOTHESIS_TESTS[args.test]()
        return

    # Run everything; tolerate missing data files
    test_order = ["H1", "H2", "H3", "H4", "H5", "H6", "bootstrap"]
    for key in test_order:
        try:
            HYPOTHESIS_TESTS[key]()
        except FileNotFoundError as e:
            print(f"\n[SKIPPED] {key}: {e}")
        except Exception as e:
            print(f"\n[ERROR] {key}: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
